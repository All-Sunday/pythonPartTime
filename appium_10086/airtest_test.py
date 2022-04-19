# @Description: TODO
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/4/7 16:59
# @File : pandos.py
# -*- encoding=utf8 -*-
# -*- encoding=utf8 -*-
__author__ = "zhoushudan"

import datetime

import openpyxl as openpyxl
from airtest.core.api import *
from airtest.cli.parser import cli_setup
import time
import logging
logger = logging.getLogger("airtest")
logger.setLevel(logging.ERROR)

def init(res_path):

    if not os.path.exists(res_path):
        wk = openpyxl.Workbook()
        ws = wk.active
        ws.append(['手机号', '活动名'])
        wk.save(res_path)
    wk = openpyxl.load_workbook(res_path)
    ws = wk.active

    if not cli_setup():
        auto_setup(__file__, logdir=True, devices=[
            "android://127.0.0.1:5037/38D4C20508015250?cap_method=MINICAP&&ori_method=MINICAPORI&&touch_method=MAXTOUCH", ])

    from poco.drivers.android.uiautomation import AndroidUiautomationPoco

    poco = AndroidUiautomationPoco(use_airtest_input=True, screenshot_each_action=False)

    # poco("com.android.systemui:id/scrim_behind")
    # script content
    print("start...")
    return poco, wk, ws


def processor(poco, res_wk, res_ws, res_path, phone_num):
    time.sleep(0.5)
    pengzhang_sign = False
    activity_sign = False
    if poco(text="营销方式").exists():
        poco(text="营销方式").swipe([-0.9332, 0.927])

    phone = poco("phone")
    # phone.click()
    phone.set_text(phone_num)
    # time.sleep(100)
    phone.click()
    time.sleep(0.2)
    # poco("phone").set_text('\n')
    keyevent("KEYCODE_SEARCH")
    time.sleep(0.5)
    try:
        poco(text="选营销案").click()
    except:
        if poco("android:id/parentPanel").exists():
            #     time.sleep(5)
            poco("android:id/button1").click()
            return
    # cover = poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child("android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[1]
    # cover = poco("android.widget.FrameLayout").offspring("android:id/inputArea").offspring("android.view.View"
    # poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child("android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[4].wait_for_appearance()
    # cover = poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child("android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[4]

    while True:
        try:
            if poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child(
                    "android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[4].exists():
                time.sleep(0.5)
        except:
            break
    if poco("android:id/parentPanel").exists():
        #     time.sleep(5)
        poco("android:id/button1").click()
        return
    poco("mkcase_fuzzy").set_text('膨胀')
    time.sleep(1)
    poco("select-fuzzy").click()
    time.sleep(2)
    if not poco("select-fuzzy").exists():
        if poco("android:id/parentPanel").exists():
            #     time.sleep(5)
            poco("android:id/button1").click()

            return

    # try:
    #     poco("select-fuzzy").click()
    # except:
    #     if poco("android:id/parentPanel").exists():
    #         #     time.sleep(5)
    #         poco("android:id/button1").click()
    #
    #         return
    poco("select-fuzzy").click()
    while True:
        try:
            if poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child(
                    "android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[4].exists():
                time.sleep(0.5)
        except:
            break

    time.sleep(1)
    # lists = poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child(
    #     "android.webkit.WebView").offspring("mkcasenode").offspring("mkcaseswiper").offspring(
    #     "android.widget.ListView").child("android.view.View")
    # print(len(lists))
    # if len(lists) == 0:
    #     return
    # for i in lists:
    #     try:
    #         #         print(i.get_text())
    #         if i.child("android.widget.RadioButton").exists():
    #             button = i.child("android.widget.RadioButton")
    #             # print(button.get_text())
    #             if ('膨胀' in button.get_text()) and ('无忧版' in button.get_text()):
    #                 button.click()
    #                 pengzhang_sign = True
    #                 break
    #     except:
    #         continue
    if poco(textMatches='.*膨胀.*无忧版.*').exists():
        poco(textMatches='.*膨胀.*无忧版.*').click()
        pengzhang_sign = True

    if not pengzhang_sign:
        return
    while True:
        try:
            if poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child(
                    "android.webkit.WebView").child("android.webkit.WebView").child("android.view.View")[4].exists():
                time.sleep(0.5)
        except:
            break

    time.sleep(1)
    # lists = poco("android.widget.FrameLayout").offspring("com.finedo.oneapp:id/ll_root").child(
    #     "android.webkit.WebView").offspring("mkcasewaynode").offspring("mkcasewayswiper").offspring(
    #     "android.widget.ListView").child("android.view.View")
    # # print(len(lists))
    # if len(lists) == 0:
    #     return
    # for i in lists:
    #     try:
    #         if i.child("android.widget.RadioButton").exists():
    #             button = i.child("android.widget.RadioButton")
    #             #
    #             if ('当月' in button.get_text()):
    #                 # print(button.get_text())
    #                 button.click()
    #                 res_ws.append([phone_num, button.get_text()])
    #     except:
    #         continue
    if poco(textMatches='.*当月.*').exists():
        button = poco(textMatches='.*当月.*')
        # button.click()
        res_ws.append([phone_num, button.get_text()])
    res_wk.save(res_path)


if __name__ == '__main__':
    file_path = input('请输入手机号xlsx路径：').strip()
    res_path = file_path.split('.')[0] + datetime.datetime.now().strftime('%Y%m%d %H-%M-%S') + 'res.xlsx'
    poco, res_wk, res_ws = init(res_path)

    source_wk = openpyxl.load_workbook(file_path)
    source_ws = source_wk.active
    rows = source_ws.max_row
    phone_data = []
    start_time = datetime.datetime.now()
    print(start_time, 'Begin-----------------------')
    for i in range(1, rows + 1):
        cell_value = source_ws.cell(row=i, column=1).value
        phone_data.append(cell_value)
        print(cell_value)
        try:
            processor(poco, res_wk, res_ws, res_path, cell_value)
        except:
            res_ws.append([cell_value])
    # print(phone_data)
    end_time = datetime.datetime.now()
    print(end_time, 'Finish-----------------------共耗时', end_time - start_time)

# generate html report
# from airtest.report.report import simple_report
# simple_report(__file__, logpath=True)