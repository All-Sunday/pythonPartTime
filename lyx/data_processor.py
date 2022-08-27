# @Description: TODO
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/8/3 11:01
# @File : data_processor.py
from openpyxl import load_workbook


def processor(path):
    res_path = path[:-5] + 'res.xlsx'
    print(res_path)

    wb = load_workbook(path)
    source_sheet = wb.active

    FRC_sheet = wb.create_sheet('FRC')
    PV_sheet = wb.create_sheet('PV')
    FV_sheet = wb.create_sheet('FV')
    RC_sheet = wb.create_sheet('RC')
    target_sheets = {'FRC': FRC_sheet, 'PV': PV_sheet, 'FV': FV_sheet, 'RC': RC_sheet}

    id = -1
    print('运行中😀')
    row_num = 1
    for row in source_sheet:
        print(row_num)
        row_num += 1
        sign_value = row[0].value
        # print('sign :', sign_value)

        sign_value1 = row[1].value
        sign_value2 = row[2].value
        if sign_value is None and sign_value1 is None and sign_value2 is None:
            continue

        if sign_value is not None:
            if isinstance(sign_value, int):
                id = sign_value
                continue

            if sign_value in target_sheets:
                target_sheet = target_sheets[sign_value]
                # print(target_sheet)
                continue

        res_row = [id]
        for cell in row[1:]:
            cell_calue = cell.value
            # print(cell_calue)
            # if cell_calue is not None:
            res_row.append(cell_calue)
        # print(res_row)
        if len(res_row) > 1:
            target_sheet.append(res_row)
            wb.save(res_path)
    print('运行结束😀')

if __name__ == '__main__':
    # path = r'数据处理(1).xlsx'
    path = input('请输入xlsx路径：').strip()
    processor(path)
