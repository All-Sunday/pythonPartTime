from decimal import Decimal
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import os
from datetime import datetime

path = r'E:\code_workplace\python\2022\january\tb_excel_filter\\'
file_list = []  # 用于保存不符合条件的文件名
target_date = (datetime.now() + relativedelta(months=1)).strftime('%Y.%m.') + 'FTE'  # 获取下一月 并转换格式'2022.04.FTE'
for parents, dirnames, filenames in os.walk(path, topdown=False):
    # 循环判断每一个文件，根据文件名
    for name in filenames:
        # 判断文件是否为表格
        if name.endswith(".xls") or name.endswith(".xlsx"):
            # 组成每个文件的绝对路径
            file = os.path.join(parents, name)

            wb = load_workbook(file)
            ws = wb.active
            # dt = datetime.today()
            # a = dt.year
            # b = dt.month

            columns = ['Project Name', 'Pool Name', 'Position Code', 'Position Name', 'Position Role', 'Work Type',
                       'Backlog Work', 'Username', target_date]  # 要保留的字段的字段名

            max_row = ws.max_row  # 最大行
            # target_cols = []  # 要保留的目标列
            for titile_cell in ws[1]:  # 循环遍历第一行表头
                # for col in range(1, max_col + 1):  # 循环遍历第一行表头
                title = titile_cell.value  # 获取表头值
                col = titile_cell.column  # 获取列号
                if title in columns:  # 判断是否是目标列
                    if title == target_date:  # 如果是日期列 则修改单元格格式
                        res = 0  # 记录总和
                        for row in range(2, max_row + 1):
                            date_cell = ws.cell(row, col)
                            res += date_cell.value  # 累加
                            date_cell.number_format = '0.0000'  # openpyxl设置单元格格式
                        res = Decimal(res).quantize(Decimal("0.0000"))  # Decimal保留四位小数
                        if res % 1 != 0:  # 判断是否是整数
                            file_list.append(name)  # 不是整数 则将文件名加入list
                    # target_cols.append(col)
                    columns.remove(title)  # 移除找到的列名  防止重复
                else:
                    ws.delete_cols(col)  # 删除不需要保留的列

            # i = 2
            # j = ws.max_row + 1
            # while i < j:  # 从第一行开始遍历
            #     i = i + 1
            #     if "2023" in str(ws.cell(row=i, column=3).value):
            #         ws.delete_rows(i)
            #         i = i - 1
            wb.save(file)

print(file_list)
