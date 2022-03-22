# @Description: 这写sheet除去前三个，分别是FB2201-FB2213,现在要改成：1、把每个已有的FB22**复制一个变成两个，且内容完全一致，这就意味着现有的FB2201变成了FB2201和FB2202,FB2202变成了FB2203和FB2204,同时，新的FB2201和FB2202和旧的FB2201在内容上完全一致，新的FB2203和FB2204和旧的FB2202在内容上完全一致,以此类推,最后应该是新的FB2225和FB2226和旧的FB2213完全一致。2、将新生成的26个sheet按照FB2201——FB2226来命名。3、要求能够批量处理。
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/3/22 11:42
# @File : copy_sheets.py
import os

from openpyxl import load_workbook


def processor(path):
    for parents, dirnames, filenames in os.walk(path, topdown=False):  # 循环判断每一个文件，根据文件名
        for name in filenames:  # 判断文件是否为表格
            if name.endswith(".xls") or name.endswith(".xlsx"):  # 组成每个文件的绝对路径
                file = os.path.join(parents, name)

                wb = load_workbook(file)
                sheets_num = len(wb.sheetnames)  # 复制sheet前的sheet数量
                offset_sheets_num = -sheets_num + 3 + 1  # 复制生成的sheet偏移量

                i = 0
                for sheet in wb:
                    i += 1
                    if i <= 3:  # 前三个sheet不处理
                        continue
                    ws_copy = wb.copy_worksheet(sheet)  # 复制sheet
                    wb.move_sheet(ws_copy, offset_sheets_num)  # 将复制生成的sheet移动到源sheet后
                    offset_sheets_num += 1  # 因复制的sheet移动 故下一个复制的sheet偏移量+1

                format_name = 'FB%d'  # 格式化sheet名称的前部
                name_num_zero = (int(wb.sheetnames[3][2:]) - 1)  # sheet名称的基数
                name_num = name_num_zero + 2 * (int(wb.sheetnames[-2][2:]) - name_num_zero)  # sheet名称的最大数
                i = 0
                sheets_num = len(wb.sheetnames)  # 复制sheet后的sheet数量
                for sheet_name in wb.sheetnames[::-1]:  # 倒序循环sheetname，修改sheet名称
                    i += 1
                    if (sheets_num - i) <= 3:  # 前三个sheet不处理
                        continue
                    wb[sheet_name].title = (format_name % name_num)  # 修改名称
                    name_num -= 1  # sheet名称后部递减

                wb.save(file)


if __name__ == '__main__':
    path = r'E:\code_workplace\python\2022\january\tb_excel_filter\cpoy_sheets\\'
    processor(path)
