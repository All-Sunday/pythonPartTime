# @Description: 判断每个sheet中ID列和PercentageSum列之间的这些列（数量不定）从第二行开始是否有为空值的单元格，只判断以FB开头的sheet，如果是空值要打印出来空值单元格所在的文件名和sheet名以及行列，并输入1后将所有空单元格自动填入“0”。还是批量处理
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/4/13 18:00
# @File : null_processor.py

from openpyxl import load_workbook
import os

from openpyxl.utils import get_column_letter

path = r'E:\code_workplace\python\2022\january\tb_excel_filter\null_processor\\'

for parents, dirnames, filenames in os.walk(path, topdown=False):  # 循环判断每一个文件，根据文件名
    for name in filenames:  # 判断文件是否为表格
        if name.endswith(".xls") or name.endswith(".xlsx"):  # 组成每个文件的绝对路径
            file = os.path.join(parents, name)

            wb = load_workbook(file)
            sheet_names = wb.sheetnames  # 一个excel文件的所有sheet名称
            # print(sheet_names)

            for sheet_name in sheet_names:  # 遍历所有sheet
                if ('FB' in sheet_name) and ('-' not in sheet_name):  # FB 开头的sheet
                    ws = wb[sheet_name]

                    max_row = ws.max_row  # 最大行
                    sign = False  # 是否是  ID列和PercentageSum列之间的这些列
                    for titile_cell in ws[2]:
                        col = titile_cell.column  # 获取列号
                        if 'ID' == titile_cell.value:  # 直到出现 ID 列 说明是目标列  sign = True
                            sign = True
                        if sign:
                            for row in range(3, max_row + 1):  # 遍历当前列的单元格
                                data_cell = ws.cell(row, col)
                                if data_cell.value is None:  # 如果是空值 输出位置
                                    print(file, sheet_name, data_cell.row, get_column_letter(data_cell.column))
                        if 'PercentageSum' == titile_cell.value:  # 直到遍历完 PercentageSum 列 说明是处理完毕  sign = False
                            sign = False

while True:
    selection = input('空单元格是否填入零（yes；no）：').strip()
    if selection in ('yes', 'no'):
        break

if selection == 'yes':  # 空值填入0
    for parents, dirnames, filenames in os.walk(path, topdown=False):  # 循环判断每一个文件，根据文件名
        for name in filenames:  # 判断文件是否为表格
            if name.endswith(".xls") or name.endswith(".xlsx"):  # 组成每个文件的绝对路径
                file = os.path.join(parents, name)

                wb = load_workbook(file)
                sheet_names = wb.sheetnames

                for sheet_name in sheet_names:
                    if ('FB' in sheet_name) and ('-' not in sheet_name):
                        ws = wb[sheet_name]

                        max_row = ws.max_row  # 最大行
                        sign = False
                        for titile_cell in ws[2]:
                            col = titile_cell.column  # 获取列号
                            if 'ID' == titile_cell.value:
                                sign = True
                            if sign:
                                for row in range(3, max_row + 1):
                                    data_cell = ws.cell(row, col)
                                    if data_cell.value is None:
                                        data_cell.value = 0  # 修改空值为0
                            if 'PercentageSum' == titile_cell.value:
                                sign = False
                wb.save(file)
