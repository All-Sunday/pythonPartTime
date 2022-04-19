import time
from decimal import Decimal
from dateutil.relativedelta import relativedelta  # 修改时间
from openpyxl import load_workbook
import os
from datetime import datetime

path = r'E:\code_workplace\python\2022\january\tb_excel_filter\新建文件夹\\'
file_list = []  # 用于保存不符合条件的文件名
next_month = (datetime.now() + relativedelta(months=1)).strftime('%Y.%m.') + 'FTE'  # 获取下一月 并转换格式'2022.04.FTE'

while True:
    selection = input('请输入选择（1：删除行列；2：保留四位小数；3：保留次月起的列）：').strip()
    if selection in ('1', '2', '3'):
        break

for parents, dirnames, filenames in os.walk(path, topdown=False):  # 循环判断每一个文件，根据文件名
    for name in filenames:  # 判断文件是否为表格
        if name.endswith(".xls") or name.endswith(".xlsx"):  # 组成每个文件的绝对路径
            file = os.path.join(parents, name)

            wb = load_workbook(file)
            ws = wb.active
            if selection == '1':  # 删除带2023的行和只保留次月列
                columns = ['Project Name', 'Pool Name', 'Position Code', 'Position Name', 'Position Role', 'Work Type',
                           'Backlog Work', 'Username', next_month]  # 要保留的列名
            elif selection == '2':  # 对所有月份保留四位小数 并输出次月加和前后不一致的表格名称
                columns = []  # 要保留的列名 空代表全部保留
            elif selection == '3':  # 保留当前年份下个月起所有月份和删除position code带2023的行
                columns = ['Project Name', 'Pool Name', 'Position Code', 'Position Name', 'Position Role', 'Work Type',
                           'Backlog Work', 'Username']
                now_year = next_month[0:4]
                start_month = int(next_month[5:7])
                for i in range(12 + 1 - start_month):
                    columns.append(now_year + '.%02d' % (start_month) + '.FTE')  # '2022.04.FTE'
                    start_month += 1

            delete_cols = []  # 要删除的列号
            max_row = ws.max_row  # 最大列

            for titile_cell in ws[1]:  # 遍历第一行表头
                title = titile_cell.value  # 获取表头内容
                col = titile_cell.column  # 获取列号
                if (title in columns) or (len(columns) == 0):  # 判断是否是目标列
                    if selection == '2':  # 对所有月份保留四位小数 并输出次月加和前后不一致的表格名称
                        if title == next_month:  # 如果是目标月份FTE列，则修改单元格格式
                            sum_before = 0  # 记录保留四位小数前的总和
                            sum_after = 0  # 记录保留四位小数后的总和
                            for row in range(2, max_row + 1):
                                date_cell = ws.cell(row, col)
                                sum_before += date_cell.value  # 累加  保留四位小数前的值
                                print(date_cell.value)
                                date_cell.value = Decimal(date_cell.value).quantize(Decimal("0.0000"))  # 保留四位小数
                                date_cell.number_format = '0.0000'  # openpyxl设置单元格格式
                                sum_after += date_cell.value  # 累加  保留四位小数后的值
                            sum_before = Decimal(sum_before).quantize(Decimal("0.0000"))  # Decimal保留四位小数
                            print(sum_before, sum_after)
                            # if sum_before % 1 != 0:  # 判断是否是整数
                            # if sum_before != sum_after:  # 判断保留四位小数前后是否相等
                            if (sum_before % 1 != 0) or (sum_before != sum_after):  # 判断是否是整数 及 保留四位小数前后是否相等
                                file_list.append(name)  # 不是整数 则将文件名加入list
                        elif 'FTE' in title:  # 如果是其他月份FTE列，保留四位小数
                            for row in range(2, max_row + 1):
                                date_cell = ws.cell(row, col)
                                date_cell.value = Decimal(date_cell.value).quantize(Decimal("0.0000"))  # 保留四位小数
                                date_cell.number_format = '0.0000'  # openpyxl设置单元格格式
                    if title == 'Position Code':  # 如果是Position Code列，则修改position_code_col值
                        position_code_col = col
                else:
                    # ws.delete_cols(col)  # 删除不需要保留的列
                    delete_cols.append(col)  # 保存删除的列号

            if (selection == '1') or (selection == '3'):  # 删除带2023的行和只保留次月列
                delete_num = 0  # 已删除列数

                for c in delete_cols:
                    ws.delete_cols(c - delete_num)  # 删除不需要保留的列
                    delete_num += 1

                i = 0
                j = ws.max_row
                while i < j:  # 从第一行开始遍历
                    i = i + 1
                    if "2023" in str(ws.cell(row=i, column=position_code_col).value):
                        ws.delete_rows(i)
                        i -= 1
            wb.save(file)
if selection == '2':  # 对所有月份保留四位小数 并输出次月加和前后不一致的表格名称
    for file in file_list:
        print(file)
