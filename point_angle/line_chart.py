# @Description: TODO
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/3/25 18:01
# @File : line_chart.py
import math
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart

if __name__ == '__main__':
    file = r'file\z-c6.1(1).xlsx'

    df = pd.read_excel(file)
    target_columns = ['EMG1.v, uV', 'EMG2.v, uV', 'EMG3.v, uV', 'EMG4.v, uV', 'EMG5.v, uV', 'EMG6.v, uV', 'EMG7.v, uV',
                      'EMG8.v, uV']
    EMG1, EMG2, EMG3, EMG4, EMG5, EMG6, EMG7, EMG8 = 0, 0, 0, 0, 0, 0, 0, 0
    EMGs = [EMG1, EMG2, EMG3, EMG4, EMG5, EMG6, EMG7, EMG8]


    for row in df.itertuples():
        for i in range(2, 10):
            EMGs[i - 2] += (row[i] * row[i])
    row_num = 0
    for EMG in EMGs:
        RMS = math.sqrt(EMG / len(df))
        df.loc[row_num, 'RMS'] = RMS
        row_num += 1
    df.to_excel(file, index=None)
    time.sleep(2)

    wb = load_workbook(file)
    ws = wb.active

    chart = LineChart()
    chart.width = 60
    chart.height = 10
    chart.x_axis.title = ws.cell(1, 1).value
    data = Reference(ws, min_col=2, min_row=1, max_col=9, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    times = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.set_categories(times)
    ws.add_chart(chart, "M6")

    for i in range(2, ws.max_column - 1):
        chart = LineChart()
        chart.width = 40
        chart.height = 5
        chart.x_axis.title = ws.cell(1, 1).value
        data = Reference(ws, min_col=i, min_row=1, max_col=i, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        times = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.set_categories(times)
        ws.add_chart(chart, "M" + str((i + 1) * 10))

    wb.save(file[:-5] + 'res.xlsx')



