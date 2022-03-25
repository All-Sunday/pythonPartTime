# @Description: TODO
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/3/25 18:01
# @File : line_chart.py
import math
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.series import Series

if __name__ == '__main__':
    file = r'file\z-c6.1(1).xlsx'

    wb = load_workbook(file)
    ws = wb.active
    for i in range(2, ws.max_column):
        chart = LineChart()  # 图表对象
        chart.width = 40
        chart.height = 5
        chart.x_axis.title = ws.cell(1, 1).value
        # chart.x_axis.crossAx = 150
        data = Reference(ws, min_col=i, min_row=1, max_col=i, max_row=ws.max_row)  # 涉及数据
        chart.add_data(data, titles_from_data=True)
        times = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.set_categories(times)
        ws.add_chart(chart, "M" + str((i - 1) * 10))  # 将图表添加到 sheet中

    wb.save(file[:-5] + 'res.xlsx')
