# @Description: requests 不行 header的 3bffc44f0adf514e9a57的值每次变
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/6/5 20:01
# @File : qcc.py
import json
import os
import time

import openpyxl
import requests
from fake_useragent import UserAgent


def init(res_path):
    os.remove(res_path)
    if not os.path.exists(res_path):
        wk = openpyxl.Workbook()
        ws1 = wk.create_sheet('省级', 0)
        ws2 = wk.create_sheet('市级', 1)
        ws1.append(
            ['	', '总计', '农、林、牧、渔业', '采矿业', '制造业', '电力、热力、燃气及水生产和供应业', '建筑业', '批发和零售业', '交通运输、仓储和邮政业', '住宿和餐饮业',
             '信息传输、软件和信息技术服务业', '金融业', '房地产业', '租赁和商务服务业', '科学研究和技术服务业', '水利、环境和公共设施管理业', '居民服务、修理和其他服务业', '教育',
             '卫生和社会工作', '文化、体育和娱乐业', '公共管理、社会保障和社会组织', '国际组织'
             ])
        ws2.append(
            ['城市', '省份', '总计', '农、林、牧、渔业', '采矿业', '制造业', '电力、热力、燃气及水生产和供应业', '建筑业', '批发和零售业', '交通运输、仓储和邮政业', '住宿和餐饮业',
             '信息传输、软件和信息技术服务业', '金融业', '房地产业', '租赁和商务服务业', '科学研究和技术服务业', '水利、环境和公共设施管理业', '居民服务、修理和其他服务业', '教育',
             '卫生和社会工作', '文化、体育和娱乐业', '公共管理、社会保障和社会组织', '国际组织'
             ])
        wk.save(res_path)
    wk = openpyxl.load_workbook(res_path)
    return wk, ws1, ws2


def query_qcc_res(header, province, industry_list, ua, wk, ws1, ws2, city={}):
    api_url = 'https://www.qcc.com/api/search/searchCount'
    if city:
        res = [city['name'], '']
        filter_json_str = "{\"r\":[{\"pr\":\"" + province['code'] + "\",\"cc\":[" + str(city) + "]}]}"
    else:
        res = [province['name']]
        filter_json_str = "{\"r\":[{\"pr\":\"" + province['code'] + "\"}]}"
    print(filter_json_str)
    for industry in industry_list:
        filter_json_str2 = filter_json_str
        if industry != '':
            # header['3bffc44f0adf514e9a57'] = 'ab4ab68059d3301fd84cdbe238a6687d8217d409688eb67709225c10ec64a6793af09efc096516089df092b7f213c1ee5a925293355b2ef06cf2ebb987b9fbd6'
            header['3bffc44f0adf514e9a57'] = '582466d5e5e629dadfd72b28e189497780bf82f0c87352ebdc102eb884266bbd624a733c7b46eff8102b4d954c1e7c5ea6c9c764e958518a17c4bddb4a403f72'  # 一个省份 一个行业
            filter_json_str2 = filter_json_str[:-1] + ',\"i\":[\"' + industry + '\"]}'
        HTTPJSONKEY = {"count": True, "filter": filter_json_str2}
        print(HTTPJSONKEY)
        # time.sleep(100)

        # HTTPJSONKEY = {"count": True, "filter": "{\"r\":[{\"pr\":\"GD\",\"cc\":[445300]}],\"i\":[\"A\"]}"}
        # HTTPJSONKEY = {"count":True,"filter":"{\"r\":[{\"pr\":\"GD\"}]}"}
        # print(HTTPJSONKEY)
        # try:
        header['User-Agent'] = ua.random
        time.sleep(0.5)
        # 关闭抓包工具
        res_json = json.loads(requests.post(api_url, headers=header, json=HTTPJSONKEY).text)
        # print(res_json)
        # except (Exception, BaseException) as e:
        #     print('qcc', e)
        #     return False, 'qcc错误' + str(e)
        # if res_json['Status'] != 200:
        if 'status' in res_json:
            print(res_json)
            return False, res_json
        print(res_json)
        print(res_json['Result']['Count'])
        res.append(res_json['Result']['Count'])
        # time.sleep(100)

    if not city:
        ws2.append(res)
    else:
        ws1.append(res)

def main():
    city_json_file = r'city.json'
    industry_json_file = r'industry.json'
    ua = UserAgent(path=r"fake_useragent.json")
    res_path ='res.xlsx'
    with open(city_json_file, encoding='utf-8') as f:
        city_json = json.loads(f.read())
    # print(city_json)
    # print(city_json[0]['name'])
    with open(industry_json_file, encoding='utf-8') as f:
        industry_json = json.loads(f.read())

    industry_list = ['']
    for industry in industry_json:
        industry_list.append(industry['code'])
    # print(industry_list)
    # time.sleep(100)

    wk, ws1, ws2 = init(res_path)
    # time.sleep(100)
    headers_qcc = {'Host': 'www.qcc.com',
                   'Connection': 'keep-alive',
                   # 'Connection': 'close',
                   'x-pid': '5fa6386a4f9af2ea1cc5ca9fd1d4bec2',
                   'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="102", "Google Chrome";v="102"',
                   'sec-ch-ua-mobile': '?0',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.63 Safari/537.36',
                   'Content-Type': 'application/json',
                   '3bffc44f0adf514e9a57': '6ff00addaaccc30bdb01777a278e2b714a9a5a29b111d180ac56fa1de22de5005943f1ae2d46d9afa660a1f54dd8e45e3d1f6665a937b59977f206fd6bd5cc20',
                   'Accept': 'application/json, text/plain, */*',
                   'X-Requested-With': 'XMLHttpRequest',
                   'sec-ch-ua-platform': "Windows",
                   'Origin': 'https://www.qcc.com',
                   'Sec-Fetch-Site': 'same-origin',
                   'Sec-Fetch-Mode': 'cors',
                   'Sec-Fetch-Dest': 'empty',
                   'Referer': 'https://www.qcc.com/web/search/advance?hasState=true',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,ru;q=0.7',
                   'Cookie': 'QCCSESSID=7848ef1e96641ae5b09ceda0fd; qcc_did=2364f334-ad95-4ba9-9a1b-de528af93d0d; UM_distinctid=18131f676923ee-0338176b7debf4-15373079-144000-18131f6769382d; acw_tc=b676109816544896370297061ee74fa452372c1d00c812a04381c4d5f3; CNZZDATA1254842228=55014950-1654400415-%7C1654486810'
                   }

    # for city in city_json:
    for province in city_json:
        # print(city['code'], city['name'])
        print(province['code'], province['name'])
        query_qcc_res(headers_qcc, province, industry_list, ua, wk, ws1, ws2)
        # if 'list' in province:
        #     for city in province['list']:
        #         print(city['code'], city['name'])
        #         query_qcc_res(headers_qcc, province, industry_list, ua, wk, ws1, ws2, city)
        # if 'list' in city:
        #     for county in city['list']:
        #         print(county['code'], county['name'])
        wk.save(res_path)
    time.sleep(100)


main()
