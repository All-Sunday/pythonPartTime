# @Description: TODO
# @Author https://github.com/All-Sunday All-Sunday
# @Time 2022/6/5 20:01
# @File : qcc.py
import json
import os
import time

import openpyxl
import requests
from fake_useragent import UserAgent
import datetime
import time

from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium import webdriver


def init(res_path):
    if os.path.exists(res_path):
        os.remove(res_path)
    if not os.path.exists(res_path):
        wk = openpyxl.Workbook()
        ws1 = wk.create_sheet('省级', 0)
        ws2 = wk.create_sheet('市级', 1)
        ws1.append(
            ['	', '总计', '农、林、牧、渔业', '采矿业', '制造业', '电力、热力、燃气及水生产和供应业', '建筑业', '批发和零售业', '交通运输、仓储和邮政业', '住宿和餐饮业',
             '信息传输、软件和信息技术服务业', '金融业', '房地产业', '租赁和商务服务业', '科学研究和技术服务业', '水利、环境和公共设施管理业', '居民服务、修理和其他服务业', '教育',
             '卫生和社会工作', '文化、体育和娱乐业', '公共管理、社会保障和社会组织', '国际组织'
             ])
        ws2.append(
            ['城市', '省份', '总计', '农、林、牧、渔业', '采矿业', '制造业', '电力、热力、燃气及水生产和供应业', '建筑业', '批发和零售业', '交通运输、仓储和邮政业', '住宿和餐饮业',
             '信息传输、软件和信息技术服务业', '金融业', '房地产业', '租赁和商务服务业', '科学研究和技术服务业', '水利、环境和公共设施管理业', '居民服务、修理和其他服务业', '教育',
             '卫生和社会工作', '文化、体育和娱乐业', '公共管理、社会保障和社会组织', '国际组织'
             ])
        wk.save(res_path)
    # wk = openpyxl.load_workbook(res_path)
    return wk, ws1, ws2


def init2(chromedriver_path, chrome_port):
    # 启动谷歌浏览器
    options = Options()
    options.add_argument('--no-sandbox')  # 解决DevToolsActivePort文件不存在的报错
    options.add_argument('window-size=1920x3000')  # 设置浏览器分辨率
    options.add_argument('--disable-gpu')  # 谷歌文档提到需要加上这个属性来规避bug
    options.add_argument('--hide-scrollbars')  # 隐藏滚动条，应对一些特殊页面
    options.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片，提升运行速度
    # options.add_argument('--headless')                  # 开启无头模式（无界面启动）

    options.add_experimental_option("debuggerAddress", "127.0.0.1:" + chrome_port)
    driver = webdriver.Chrome(chromedriver_path, options=options)
    # driver.execute_script('window.open("' + start_url + '")')
    # driver.switch_to.window(driver.window_handles[-1])
    # time.sleep(0.5)
    # driver.switch_to.window(driver.window_handles[0])

    return driver


def query_qcc_res(wk, ws1, ws2, source_path, driver, target_cities):
    res_province = []
    res_city = []
    print('sssss')
    # driver.find_element_by_xpath('//div[@class="filter-area"][1]/div[2]/div[2]/div/div/span').click()  # 全部行业  出现下拉选择
    # time.sleep(0.2)
    # driver.find_element_by_xpath(
    #     '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/div/div/div/div/ul/li[1]/a/label/input').click()  #循环点击一个行业
    # time.sleep(0.1)
    # driver.execute_script("arguments[0].click();", driver.find_element_by_xpath(
    #     '/html/body/div[1]/div[2]/div[1]/div/div[1]/div/div[2]/div[3]')) # 点击常用筛选 隐藏下拉框
    # ActionChains(driver).move_to_element(driver.find_element_by_xpath(
    #     '/html/body/div[1]/div[2]/div[1]/div/div[1]/div/div[2]/div[3]')).perform()
    # driver.find_element_by_xpath(
    #     '//input[@class="form-control searchkey"]').click()  # 点击输入关键词 隐藏下拉框
    # 点击常用筛选 隐藏下拉框
    time.sleep(0.1)
    driver.find_element_by_xpath('//div[@class="filter-area"][1]/div[4]/div[2]/div/div/span').click()  # 全部地区  出现下拉选择省份
    time.sleep(0.2)
    provinces = driver.find_elements_by_xpath(
        '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/div/div/div/div/ul/li')  # 全部省份
    provinces_length = len(provinces) - 3
    # for province in provinces:
    # for province_index in range(15, provinces_length):
    for province_index in range(28, 31):
        # 无需取消勾选  已经操作过
        # if province_index != 0:
        #     driver.execute_script("arguments[0].click();",
        #                           provinces[province_index - 1].find_element_by_xpath('./a/label/input'))  # 取消勾选上一个省份
        province = provinces[province_index]
        driver.execute_script("arguments[0].click();", province.find_element_by_xpath('./a/label/input'))  # 循环点击省份
        province_name = province.find_element_by_xpath('./a').get_attribute('title')  # 省份名称
        time.sleep(1.6)

        province_total_num = eval(driver.find_element_by_xpath('//span[@class="smultitext"]/span').text.strip())  # 全部行业
        # res_province.append([province_name, province_total_num])
        res_province_single = [province_name, province_total_num]
        # print(res_province)

        driver.find_element_by_xpath(
            '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/span/a/span[2]').click()  # 全部行业  出现下拉选择
        time.sleep(0.1)
        industries = driver.find_elements_by_xpath(
            '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/div/div/div/div/ul/li')
        # industries_length = len(industries)
        industries_length = 21
        for industry_index in range(industries_length):
            # print(industries_length, industry_index)
            if industry_index != 0:
                driver.execute_script("arguments[0].click();", industries[industry_index - 1].find_element_by_xpath(
                    './a/label/input'))  # 取消勾选上一个行业
            if industry_index == (industries_length - 1):
                break
            # print(industries_length, industry_index)
            driver.execute_script("arguments[0].click();",
                                  industries[industry_index].find_element_by_xpath('./a/label/input'))
            time.sleep(1.6)
            province_num = eval(
                driver.find_element_by_xpath('//span[@class="smultitext"]/span').text.strip())  # 全部行业
            res_province_single.append(province_num)
            # print(res_province_single)
        res_province.append(res_province_single)
        print(len(res_province_single))
        ws1.append(res_province_single)
        wk.save(source_path)
        # print(res_province)
        driver.find_element_by_xpath(
            '//input[@class="form-control searchkey"]').click()  # 点击输入关键词 隐藏下拉框
        time.sleep(0.3)

        # time.sleep(100)

        driver.find_element_by_xpath(
            '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/span/a').click()  # 全部地区  出现下拉选择省份
        time.sleep(0.2)
        ActionChains(driver).move_to_element(provinces[province_index + 1]).perform()
        time.sleep(0.2)
        ActionChains(driver).move_to_element(province).perform()
        time.sleep(1.6)
        citys = driver.find_elements_by_xpath(
            '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/div/div/div[2]/div/ul/li')  # 全部城市
        for city in citys:
            driver.execute_script("arguments[0].click();", city.find_element_by_xpath('./a/label/input'))  # 循环取消勾选全部城市
            # city.find_element_by_xpath('./a/label/input').click()  # 循环点击城市
            time.sleep(0.1)
        citys_length = len(citys)
        for city_index in range(citys_length):
            if city_index != 0:
                driver.find_element_by_xpath(
                    '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/span/a').click()  # 全部地区  出现下拉选择省份

                # driver.find_element_by_xpath(
                #     '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/span/a/span[2]').click()  # 全部地区  出现下拉选择省份
                time.sleep(0.2)
                ActionChains(driver).move_to_element(provinces[province_index + 1]).perform()
                time.sleep(0.3)
                ActionChains(driver).move_to_element(province).perform()
                time.sleep(1.8)
                # time.sleep(0.5)
                # todo
                citys = driver.find_elements_by_xpath(
                    '//div[@class="filter-area"][1]/div[4]/div[2]/div/div/div/div/div[2]/div/ul/li')  # 全部城市
                # print(len(citys))
                # print(citys)
                # driver.execute_script("arguments[0].click();", province.find_element_by_xpath('./a/label/span'))  # 取消勾选省份
                # driver.execute_script("arguments[0].click();", province.find_element_by_xpath('./a/label/span'))  # 勾选省份
                # for city in citys:
                #     driver.execute_script("arguments[0].click();",
                #                           city.find_element_by_xpath('./a/label/input'))  # 循环取消勾选全部城市
                #     # city.find_element_by_xpath('./a/label/input').click()  # 循环点击城市
                #     time.sleep(0.1)

                # time.sleep(100)

                driver.execute_script("arguments[0].click();",
                                      citys[city_index - 1].find_element_by_xpath('./a/label/input'))  # 取消勾选上一个城市
            driver.execute_script("arguments[0].click();",
                                  citys[city_index].find_element_by_xpath('./a/label/input'))  # 循环勾选一个城市
            city_name = citys[city_index].find_element_by_xpath('./a').get_attribute('title')  # 城市名称
            if city_name not in target_cities:
                continue
            # time.sleep(0.15)
            time.sleep(1.6)

            city_total_num = eval(
                driver.find_element_by_xpath('//span[@class="smultitext"]/span').text.strip())  # 全部行业
            res_city_single = [city_name, city_total_num]

            driver.find_element_by_xpath(
                '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/span/a/span[2]').click()  # 全部行业  出现下拉选择
            time.sleep(0.2)
            industries = driver.find_elements_by_xpath(
                '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/div/div/div/div/ul/li')
            # industries_length = len(industries)
            industries_length = 21
            for industry_index in range(industries_length):
                # print(industries_length, industry_index)
                if industry_index != 0:
                    driver.execute_script("arguments[0].click();", industries[industry_index - 1].find_element_by_xpath(
                        './a/label/input'))  # 取消勾选上一个行业
                    # industries[industry_index - 1].find_element_by_xpath('./a/label/input').click() # 取消勾选上一个行业
                if industry_index == (industries_length - 1):
                    break
                driver.execute_script("arguments[0].click();",
                                      industries[industry_index].find_element_by_xpath('./a/label/input'))
                # print(industries[industry_index].find_element_by_xpath('./a').get_attribute('title'))
                time.sleep(1.6)
                city_num = eval(
                    driver.find_element_by_xpath('//span[@class="smultitext"]/span').text.strip())  # 单个行业
                res_city_single.append(city_num)
            res_city.append(res_city_single)
            print(len(res_city_single))
            ws2.append(res_city_single)
            wk.save(source_path)
            # driver.find_element_by_xpath(
            #     '//div[@class="filter-area"][1]/div[2]/div[2]/div/div/div/div/div/div/ul/li[1]/a/label/input').click()  # 循环点击一个行业
            driver.find_element_by_xpath(
                '//input[@class="form-control searchkey"]').click()  # 点击输入关键词 隐藏下拉框
            time.sleep(0.3)
            # print(res_city)

        filter_a_list = driver.find_elements_by_xpath('//div[@class="pills-after yxtj"]/a')
        for i in range(len(filter_a_list)):
            if i == 0:
                continue
            filter_a_list[i].click()
        time.sleep(0.2)
        time.sleep(1)

    print('1111')

    # filter_a_list = driver.find_elements_by_xpath('//div[@class="pills-after yxtj"]/a')
    # for i in range(len(filter_a_list)):
    #     if i == 0:
    #         continue
    #     filter_a_list[i].click()
    # time.sleep(0.2)
    time.sleep(100)


def main():
    source_path = 'res2.xlsx'
    wk, ws1, ws2 = init(source_path)
    # domain = ['https://www.ic.net.cn']

    target_cities = ['安庆市', '蚌埠市', '亳州市', '池州市', '滁州市', '阜阳市', '合肥市', '淮北市', '淮南市', '黄山市', '六安市', '马鞍山市', '铜陵市', '芜湖市',
                     '宿州市', '宣城市', '北京市', '福州市', '龙岩市', '南平市', '宁德市', '莆田市', '泉州市', '三明市', '厦门市', '漳州市', '白银市', '定西市',
                     '甘南藏族自治州', '嘉峪关市', '金昌市', '酒泉市', '兰州市', '临夏回族自治州', '陇南市', '平凉市', '庆阳市', '天水市', '武威市', '张掖市', '潮州市',
                     '东莞市', '佛山市', '广州市', '河源市', '惠州市', '江门市', '揭阳市', '茂名市', '梅州市', '清远市', '汕头市', '汕尾市', '韶关市', '深圳市',
                     '阳江市', '云浮市', '湛江市', '肇庆市', '中山市', '珠海市', '百色市', '北海市', '崇左市', '防城港市', '贵港市', '桂林市', '河池市', '贺州市',
                     '来宾市', '柳州市', '南宁市', '钦州市', '梧州市', '玉林市', '安顺市', '毕节市', '贵阳市', '六盘水市', '黔东南苗族侗族自治州', '黔南布依族苗族自治州',
                     '黔西南布依族苗族自治州', '铜仁市', '遵义市', '海口市', '三沙市', '三亚市', '保定市', '沧州市', '承德市', '邯郸市', '衡水市', '廊坊市', '秦皇岛市',
                     '石家庄市', '唐山市', '邢台市', '张家口市', '安阳市', '鹤壁市', '焦作市', '开封市', '洛阳市', '漯河市', '南阳市', '平顶山市', '濮阳市',
                     '三门峡市',
                     '商丘市', '新乡市', '信阳市', '许昌市', '郑州市', '周口市', '驻马店市', '大庆市', '大兴安岭地区', '哈尔滨市', '鹤岗市', '黑河市', '鸡西市',
                     '佳木斯市',
                     '牡丹江市', '七台河市', '齐齐哈尔市', '双鸭山市', '绥化市', '伊春市', '鄂州市', '恩施土家族苗族自治州', '黄冈市', '黄石市', '荆门市', '荆州市',
                     '十堰市',
                     '随州市', '武汉市', '咸宁市', '襄阳市', '孝感市', '宜昌市', '常德市', '郴州市', '衡阳市', '怀化市', '娄底市', '邵阳市', '湘潭市',
                     '湘西土家族苗族自治州', '益阳市', '永州市', '岳阳市', '张家界市', '长沙市', '株洲市', '白城市', '白山市', '吉林市', '辽源市', '四平市', '松原市',
                     '通化市', '长春市', '常州市', '淮安市', '连云港市', '南京市', '南通市', '苏州市', '泰州市', '无锡市', '宿迁市', '徐州市', '盐城市', '扬州市',
                     '镇江市', '抚州市', '赣州市', '吉安市', '景德镇市', '九江市', '南昌市', '萍乡市', '上饶市', '新余市', '宜春市', '鹰潭市', '鞍山市', '本溪市',
                     '朝阳市', '大连市', '丹东市', '抚顺市', '阜新市', '葫芦岛市', '锦州市', '辽阳市', '盘锦市', '沈阳市', '铁岭市', '延边朝鲜族自治州', '营口市',
                     '阿拉善盟', '巴彦淖尔市', '包头市', '赤峰市', '鄂尔多斯市', '呼和浩特市', '呼伦贝尔市', '通辽市', '乌海市', '乌兰察布市', '锡林郭勒盟', '兴安盟',
                     '固原市',
                     '石嘴山市', '吴忠市', '银川市', '中卫市', '果洛藏族自治州', '海北藏族自治州', '海东市', '海南藏族自治州', '海西蒙古族藏族自治州', '黄南藏族自治州',
                     '西宁市',
                     '玉树藏族自治州', '滨州市', '德州市', '东营市', '菏泽市', '济南市', '济宁市', '聊城市', '临沂市', '青岛市', '日照市', '泰安市', '威海市',
                     '潍坊市',
                     '烟台市', '枣庄市', '淄博市', '大同市', '晋城市', '晋中市', '临汾市', '吕梁市', '朔州市', '太原市', '忻州市', '阳泉市', '运城市', '长治市',
                     '安康市', '宝鸡市', '汉中市', '商洛市', '铜川市', '渭南市', '西安市', '咸阳市', '延安市', '榆林市', '上海市', '阿坝藏族羌族自治州', '巴中市',
                     '成都市',
                     '达州市', '德阳市', '甘孜藏族自治州', '广安市', '广元市', '乐山市', '凉山彝族自治州', '泸州市', '眉山市', '绵阳市', '南充市', '内江市', '攀枝花市',
                     '遂宁市', '雅安市', '宜宾市', '资阳市', '自贡市', '天津市', '阿里地区', '昌都市', '拉萨市', '林芝市', '那曲市', '日喀则市', '山南市',
                     '阿克苏地区',
                     '阿勒泰地区', '巴音郭楞蒙古自治州', '博尔塔拉蒙古自治州', '昌吉回族自治州', '哈密市', '和田地区', '喀什地区', '克拉玛依市', '克孜勒苏柯尔克孜自治州',
                     '塔城地区',
                     '吐鲁番市', '乌鲁木齐市', '伊犁哈萨克自治州', '保山市', '楚雄彝族自治州', '大理白族自治州', '德宏傣族景颇族自治州', '迪庆藏族自治州', '红河哈尼族彝族自治州',
                     '昆明市',
                     '丽江市', '临沧市', '怒江傈僳族自治州', '普洱市', '曲靖市', '文山壮族苗族自治州', '西双版纳傣族自治州', '玉溪市', '昭通市', '杭州市', '湖州市',
                     '嘉兴市',
                     '金华市', '丽水市', '宁波市', '衢州市', '绍兴市', '台州市', '温州市', '舟山市', '重庆市']
    s_date = '20190101'
    e_date = '20210909'

    chromedriver_path = r'E:\code_workplace\python\chromedriver.exe'
    chrome_port = '9222'

    # i = 0
    # a_list_num = 0
    # start(s_date, e_date, df, key_words, file_path[:-5] + 'res.xlsx', a_list_num,  chromedriver_path,         chrome_port)
    driver = init2(chromedriver_path, chrome_port)
    query_qcc_res(wk, ws1, ws2, source_path, driver, target_cities)


main()


